import pandas as pd
from dotenv import load_dotenv
from pathlib import Path
import os
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

# Path to the LAMP_List Excel file
LAMP_LIST_PATH = os.getenv("LAMP_LIST_PATH")
if not LAMP_LIST_PATH:
    raise ValueError("LAMP_LIST_PATH environment variable not set")

def read_lamp_list():
    """Read Excel file into pandas DataFrame using openpyxl"""
    if not os.path.exists(LAMP_LIST_PATH):
        raise FileNotFoundError(f"LAMP_List not found at {LAMP_LIST_PATH}")
    
    try:
        # Load workbook using openpyxl
        wb = load_workbook(LAMP_LIST_PATH, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = [cell.value for cell in ws[1]]
        
        # Get data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Clean up column names
        df.columns = df.columns.str.strip()
        
        # Rename specific columns
        column_mapping = {
            'Unnamed: 1': 'Company Name',
            'Contact #1': 'Primary Contact',
            'Contact #2': 'Secondary Contact',
            'Contact Dates': 'Contact History',
            'Informational Interview': 'Interview Notes'
        }
        df = df.rename(columns=column_mapping)
        
        # Fill NaN values in Company Name column if it exists
        if 'Company Name' in df.columns:
            df['Company Name'] = df['Company Name'].ffill()
        
        return df
    
    except Exception as e:
        raise e
    finally:
        wb.close()

def write_to_lamp_list(df):
    """Write DataFrame to Excel file"""
    if not os.path.exists(LAMP_LIST_PATH):
        raise FileNotFoundError(f"LAMP_List not found at {LAMP_LIST_PATH}")
    
    # Create temporary file path with correct extension
    temp_path = os.path.splitext(LAMP_LIST_PATH)[0] + '_temp.xlsx'
    
    try:
        # Write to temporary file
        df.to_excel(temp_path, index=False)
        
        # Replace original file with temporary file
        import shutil
        shutil.move(temp_path, LAMP_LIST_PATH)
        
    except Exception as e:
        # Clean up temporary file if something went wrong
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise e

if __name__ == "__main__":
    try:
        # Example usage
        df = read_lamp_list()
        print("Column headers:")
        print(df.columns.tolist())
        
        # Print Motivation column if it exists
        if 'Motivation' in df.columns:
            print("\nMotivation column data:")
            print(df[['Company Name', 'Motivation']])
    except Exception as e:
        print(f"Error: {e}")